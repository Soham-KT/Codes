import copy
import openpyxl
import numpy as np


def gradient(x, y, w, b):
    m, n = x.shape
    dj_dw = np.zeros(3, float)
    dj_db = 0.

    for i in range(m):
        f_wb = (np.dot(x[i], w) + b) - y[i]
        for j in range(n):
            dj_dw += f_wb * x[i]
        dj_db += f_wb

    dj_dw /= m
    dj_db /= m

    return dj_dw, dj_db


def gradient_descent(x, y, w_ini, b_ini, alpha, iters, compute):
    w = copy.deepcopy(w_ini)
    b = b_ini

    for i in range(iters):
        dj_dw, dj_db = compute(x, y, w, b)
        w -= alpha * dj_dw
        b -= alpha * dj_db

        return w, b


if __name__ == '__main__':
    # ------------------------------------------------ Extracting data from excel
    address = "excel_sheet\\python_gradient_descent.xlsx"
    wb = openpyxl.load_workbook(address)
    sheet = wb.sheetnames

    sheet1 = wb['Sheet1']
    X_data = np.empty((0, 3), float)
    Y_data = np.empty(1, float)
    temp_list = []

    for i in range(2, 38):
        for j in range(1, 4):
            temp_list.append(sheet1.cell(i, j).value)
        temp = np.array(temp_list, float)
        temp = temp.reshape(1, 3)
        X_data = np.append(X_data, temp, axis=0)
        temp_list.clear()

    for i in range(2, 37):
        Y_data = np.append(Y_data, sheet1.cell(i, 4).value)

    # ------------------------------------------------ Gradient Descent Code
    w_initial = np.zeros(3, float)
    b_initial = 0.
    alpha = 0.01
    iters = 100
    w_final, b_final = gradient_descent(X_data, Y_data, w_initial, b_initial, alpha, iters, gradient)
    print(w_final)
    print(b_final)



